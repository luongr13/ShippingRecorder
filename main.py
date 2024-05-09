import csv
import datetime
import os
from openpyxl import load_workbook


# string format: 20230830
def get_year(string):
    try:
        return int(string[0:4])
    except TypeError:
        print('TypeError for get_year()')
        exit()


# string format: 20230830
def get_month(string):
    try:
        return int(string[4:6])
    except TypeError:
        print('TypeError for get_month()')
        exit()


# string format: 20230830
def get_day(string):
    try:
        return int(string[6:])
    except TypeError:
        print('TypeError for get_day()')
        exit()


def get_ticket_num(string):
    if string.isdigit():
        return int(string)
    else:
        return string


# format_attn(attn, val)
# attn - type list
# val - type integer
#
# Given a list that contains the words from ATTN field,
# return a string with first name abbreviated or full name.
def format_attn(to_attn, from_attn, pack_type, ref5):
    # is requestor to_attn or from_attn?
    if pack_type[0] == 'd':
        attn = from_attn.split()
    else:
        attn = to_attn.split()

    # get the name
    try:
        cutoff = attn.index('-')
    except ValueError:
        cutoff = len(attn)
   
    attn = attn[0:cutoff]
    
    ret = ''
    if pack_type[1] == '0':
        ret += attn[0][0] + '. ' + attn[1]
    elif pack_type[1] == '1':
        for i in range(len(attn)):
            ret += attn[i] + ' '
        ret = ret.rstrip()

    # concat ref5 if it is an ebay sale
    if pack_type[0] == 'c':
        ret += ': ' + ref5
    
    return ret


# Given the string the contains the company
# name in the shipping label,
# return a list of valid keywords.
def get_keywords(string):
    key_words = string.split()
    invalid = ['of',
               'the',
               'a',
               'at',
               ',',
               'healthcare',
               'health',
               'rehab',
               'rehabilitation',
               'center',
               '&',
               'nursing',
               'and',
               'care']
    for bad_key in invalid:
        while key_words.count(bad_key) > 0:
            key_words.remove(bad_key)
    return key_words


# Load the CustomerInfo
facilities = dict()
with open('dQBCustomerInfo.csv', mode='r', encoding='cp1252') as file:
    csvFile = csv.reader(file)
    next(csvFile)
    
    line_num = 1
    for line in csvFile:
        if line[3].lower() in facilities:
            print('Found duplicate: line #' + str(line_num))
            break
        elif line[3] != '':
            facilities[line[3].lower()] = line[0]
        line_num += 1


# Load the db
#
# Key: Tracking Number
# Value: List of shipment information
db = dict()
with open('dump.csv', mode='r') as file:
    csvFile = csv.reader(file)
    for v in csvFile:
        if v[0] == 'Y':
            db.pop(v[5])
        else:
            db[v[5]] = v


# Load the Workbook
workbook1_filename = 'choinventory.xlsm'
workbook2_filename = 'template.xlsm'

try:
    book1 = load_workbook(filename=workbook1_filename, read_only=False, keep_vba=True)
except FileNotFoundError:
    print('File ' + workbook1_filename + ' does not exist.')
    print('Terminating...')
    exit()

try:
    book2 = load_workbook(filename=workbook2_filename, read_only=False, keep_vba=True)
except FileNotFoundError:
    print('File ' + workbook2_filename + ' does not exist.')
    print('Terminating...')
    exit()

sheet = book1.active
sheet2 = book2.active

current_line = sheet['A2'].value
current_line2 = sheet2['A2'].value

formula = '=IFERROR(INDEX(dQBCustomerInfo[FacilityName],MATCH(Table1[[#This Row],[Facility]],dQBCustomerInfo[ShortName],0)),"")'

#   A  |  B |   C  |  D  |     E    |   F  |       G      |  H  |   I  |   J   |    K   |     L    |   M  |     N
# Tech | CC | Date | Tkt | Facility | Attn |Real Facility | Qty | Cost | Quote | Vendor | Tracking | Item | Ship to
for k in db:
    shipment = db[k]

    # Technician
    sheet.cell(current_line, 1).value = "Richard"
    sheet2.cell(current_line2, 1).value = "Richard"

    # CC
    sheet.cell(current_line, 2).value = 4812
    sheet2.cell(current_line2, 2).value = 4812

    # Date
    sheet.cell(current_line, 3).value = datetime.datetime(get_year(shipment[13]),
                                                          get_month(shipment[13]),
                                                          get_day(shipment[13]))
    sheet2.cell(current_line2, 3).value = datetime.datetime(get_year(shipment[13]),
                                                            get_month(shipment[13]),
                                                            get_day(shipment[13]))

    # Ticket Number
    if shipment[6].isdigit():
        sheet.cell(current_line, 4).value = int(shipment[6])
        sheet2.cell(current_line2, 4).value = int(shipment[6])

    else:
        sheet.cell(current_line, 4).value = shipment[6]
        sheet2.cell(current_line2, 4).value = shipment[6]

    # Facility Short Hand
    if shipment[7][0] == 'c':
        facility_name = None
    elif shipment[7][0] == 'd':
        facility_name = shipment[3]
    else:
        facility_name = shipment[1]

    if facility_name is not None:   # Attempt to search for facility shorthand name
        keywords = get_keywords(facility_name.lower())
        if shipment[8] != "":
            keywords.append(shipment[8].lower())

        for j in facilities:
            if all([x in j for x in keywords]):
                sheet.cell(current_line, 5).value = facilities[j]
                sheet2.cell(current_line2, 5).value = facilities[j]
                break
        if sheet.cell(current_line, 5).value is None:
            sheet.cell(current_line, 5).value = facility_name
            sheet2.cell(current_line2, 5).value = facility_name

    # Requestor
    sheet.cell(current_line, 6).value = format_attn(shipment[2],
                                                    shipment[4],
                                                    shipment[7],
                                                    shipment[9])
    sheet2.cell(current_line2, 6).value = format_attn(shipment[2],
                                                      shipment[4],
                                                      shipment[7],
                                                      shipment[9])

    # Real Facility
    sheet.cell(current_line, 7).value = formula
    sheet2.cell(current_line2, 7).value = formula

    # Quantity
    sheet.cell(current_line, 8).value = 1
    sheet2.cell(current_line2, 8).value = 1

    # Cost
    sheet.cell(current_line, 9).value = float(shipment[12])
    sheet2.cell(current_line2, 9).value = float(shipment[12])

    # Vendor
    sheet.cell(current_line, 11).value = 'UPS'
    sheet2.cell(current_line2, 11).value = 'UPS'

    # Order / Tracking #
    sheet.cell(current_line, 12).value = shipment[5]
    sheet2.cell(current_line2, 12).value = shipment[5]

    # Item Description
    sheet.cell(current_line, 13).value = "Shipping " + shipment[11]
    sheet2.cell(current_line2, 13).value = "Shipping " + shipment[11]

    # Ship to
    if shipment[7][0] == 'd':
        sheet.cell(current_line, 14).value = 'CTS'
        sheet2.cell(current_line2, 14).value = 'CTS'
    else:
        sheet.cell(current_line, 14).value = 'Facility'
        sheet2.cell(current_line2, 14).value = 'Facility'

    current_line += 1
    current_line2 += 1

# Move to next line number to set up for next run
sheet['A2'].value = current_line + 1
sheet2['A2'].value = current_line2 + 1

# Save to files and create backups
book1.save('choinventory.xlsm')
book2.save('todaysinventory.xlsm')
os.remove('csvbackup.csv')
os.rename('dump.csv', 'csvbackup.csv')
